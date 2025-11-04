from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from fastapi.responses import FileResponse
from auth import get_db, get_current_user
import models
import openpyxl
import os
import csv

router = APIRouter(prefix="/reportes", tags=["Reportes"])


# ✅ Exportar reporte de clientes a Excel
@router.get("/clientes/excel")
def exportar_clientes_excel(db: Session = Depends(get_db), user=Depends(get_current_user)):

    clientes = db.query(models.Cliente).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Encabezados
    ws.append([
        "ID", "Nombre", "Cédula", "Teléfono", "Correo", "Dirección",
        "Monto", "Fecha", "Estado"
    ])

    # Registros
    for c in clientes:
        ws.append([
            c.id, c.nombre, c.cedula, c.telefono, c.correo,
            c.direccion, c.monto, str(c.fecha), c.estado
        ])

    # Guardar archivo temporal
    archivo = "clientes_reporte.xlsx"
    wb.save(archivo)

    return FileResponse(
        path=archivo,
        filename=archivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ✅ Exportar reporte de préstamos a Excel
@router.get("/prestamos/excel")
def exportar_prestamos_excel(db: Session = Depends(get_db), user=Depends(get_current_user)):

    prestamos = db.query(models.Prestamo).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prestamos"

    ws.append([
        "ID", "Cliente", "Monto Total", "Pagado", "Restante",
        "Fecha Inicio", "Fecha Límite", "Estado"
    ])

    for p in prestamos:
        ws.append([
            p.id,
            p.cliente.nombre if p.cliente else "Sin cliente",
            p.monto_inicial + p.total_interes,
            p.monto_pagado,
            p.monto_restante,
            str(p.fecha_inicio),
            str(p.fecha_limite),
            p.estado
        ])

    archivo = "prestamos_reporte.xlsx"
    wb.save(archivo)

    return FileResponse(
        path=archivo,
        filename=archivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ✅ Exportar reporte de pagos a Excel
@router.get("/pagos/excel")
def exportar_pagos_excel(db: Session = Depends(get_db), user=Depends(get_current_user)):

    pagos = db.query(models.Pago).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"

    ws.append(["ID Pago", "ID Préstamo", "Monto", "Fecha", "Estado"])

    for p in pagos:
        ws.append([
            p.id, p.prestamo_id, p.monto_pagado,
            str(p.fecha_pago), p.estado
        ])

    archivo = "pagos_reporte.xlsx"
    wb.save(archivo)

    return FileResponse(
        path=archivo,
        filename=archivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# ✅ Exportar clientes CSV
@router.get("/clientes/csv")
def exportar_clientes_csv(db: Session = Depends(get_db), user=Depends(get_current_user)):

    clientes = db.query(models.Cliente).all()
    archivo = "clientes_reporte.csv"

    with open(archivo, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Nombre", "Cédula", "Teléfono", "Correo", "Dirección", "Monto", "Fecha", "Estado"])

        for c in clientes:
            writer.writerow([
                c.id, c.nombre, c.cedula, c.telefono, c.correo,
                c.direccion, c.monto, str(c.fecha), c.estado
            ])

    return FileResponse(
        archivo,
        media_type="text/csv",
        filename=archivo
    )


# ✅ Exportar préstamos CSV
@router.get("/prestamos/csv")
def exportar_prestamos_csv(db: Session = Depends(get_db), user=Depends(get_current_user)):

    prestamos = db.query(models.Prestamo).all()
    archivo = "prestamos_reporte.csv"

    with open(archivo, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Cliente", "Monto Total", "Pagado", "Restante", "Fecha Inicio", "Fecha Límite", "Estado"])

        for p in prestamos:
            writer.writerow([
                p.id,
                p.cliente.nombre if p.cliente else "Sin cliente",
                p.monto_inicial + p.total_interes,
                p.monto_pagado,
                p.monto_restante,
                str(p.fecha_inicio),
                str(p.fecha_limite),
                p.estado
            ])

    return FileResponse(
        archivo,
        media_type="text/csv",
        filename=archivo
    )


# ✅ Exportar pagos CSV
@router.get("/pagos/csv")
def exportar_pagos_csv(db: Session = Depends(get_db), user=Depends(get_current_user)):

    pagos = db.query(models.Pago).all()
    archivo = "pagos_reporte.csv"

    with open(archivo, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID Pago", "ID Préstamo", "Monto", "Fecha", "Estado"])

        for p in pagos:
            writer.writerow([
                p.id,
                p.prestamo_id,
                p.monto_pagado,
                str(p.fecha_pago),
                p.estado
            ])

    return FileResponse(
        archivo,
        media_type="text/csv",
        filename=archivo
    )


from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from fastapi.responses import FileResponse


# ✅ Exportar clientes a PDF
@router.get("/clientes/pdf")
def clientes_pdf(db: Session = Depends(get_db), user=Depends(get_current_user)):

    archivo = "clientes_reporte.pdf"
    clientes = db.query(models.Cliente).all()

    # PDF
    pdf = SimpleDocTemplate(archivo, pagesize=letter)
    elementos = []
    estilos = getSampleStyleSheet()

    elementos.append(Paragraph("REPORTE DE CLIENTES", estilos["Title"]))

    tabla_data = [["ID", "Nombre", "Cédula", "Teléfono", "Correo", "Monto", "Estado"]]

    for c in clientes:
        tabla_data.append([
            c.id, c.nombre, c.cedula, c.telefono, c.correo, c.monto, c.estado
        ])

    tabla = Table(tabla_data)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold")
    ]))

    elementos.append(tabla)
    pdf.build(elementos)

    return FileResponse(archivo, media_type="application/pdf", filename=archivo)


# ✅ Exportar préstamos a PDF
@router.get("/prestamos/pdf")
def prestamos_pdf(db: Session = Depends(get_db), user=Depends(get_current_user)):

    archivo = "prestamos_reporte.pdf"
    prestamos = db.query(models.Prestamo).all()

    pdf = SimpleDocTemplate(archivo, pagesize=letter)
    elementos = []
    estilos = getSampleStyleSheet()

    elementos.append(Paragraph("REPORTE DE PRÉSTAMOS", estilos["Title"]))

    tabla_data = [["ID", "Cliente", "Monto Total", "Pagado", "Restante", "Estado"]]

    for p in prestamos:
        tabla_data.append([
            p.id,
            p.cliente.nombre if p.cliente else "Sin cliente",
            p.monto_inicial + p.total_interes,
            p.monto_pagado,
            p.monto_restante,
            p.estado
        ])

    tabla = Table(tabla_data)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold")
    ]))

    elementos.append(tabla)
    pdf.build(elementos)

    return FileResponse(archivo, media_type="application/pdf", filename=archivo)


# ✅ Exportar pagos a PDF
@router.get("/pagos/pdf")
def pagos_pdf(db: Session = Depends(get_db), user=Depends(get_current_user)):

    archivo = "pagos_reporte.pdf"
    pagos = db.query(models.Pago).all()

    pdf = SimpleDocTemplate(archivo, pagesize=letter)
    elementos = []
    estilos = getSampleStyleSheet()

    elementos.append(Paragraph("REPORTE DE PAGOS", estilos["Title"]))

    tabla_data = [["ID Pago", "Préstamo", "Monto", "Fecha", "Estado"]]

    for p in pagos:
        tabla_data.append([
            p.id,
            p.prestamo_id,
            p.monto_pagado,
            str(p.fecha_pago),
            p.estado
        ])

    tabla = Table(tabla_data)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold")
    ]))

    elementos.append(tabla)
    pdf.build(elementos)

    return FileResponse(archivo, media_type="application/pdf", filename=archivo)
